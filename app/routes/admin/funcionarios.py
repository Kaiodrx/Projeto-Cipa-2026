import io
from datetime import date as _date
from flask import render_template, request, redirect, url_for, flash, send_file
from openpyxl import load_workbook, Workbook
from . import admin_bp
from ...extensions import db
from ...models import Funcionario
from ...constants import UNIDADES
from ...utils import login_required


def _parse_date(value):
    """Converte string 'YYYY-MM-DD' ou objeto date para date; retorna None se inválido."""
    if not value:
        return None
    if isinstance(value, _date):
        return value
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return _date.fromisoformat(str(value).strip()) if fmt == '%Y-%m-%d' else \
                   _date(*[int(p) for p in reversed(str(value).strip().split(fmt[1]))])
        except Exception:
            pass
    try:
        return _date.fromisoformat(str(value).strip())
    except Exception:
        return None


@admin_bp.route('/funcionarios')
@login_required
def funcionarios():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '')
    unidade_filtro = request.args.get('unidade', '')
    ativo_filtro = request.args.get('ativo', '')
    page = request.args.get('page', 1, type=int)

    query = Funcionario.query
    if search:
        query = query.filter(
            db.or_(
                Funcionario.nome.ilike(f'%{search}%'),
                Funcionario.matricula.ilike(f'%{search}%'),
                Funcionario.setor.ilike(f'%{search}%'),
                Funcionario.cargo.ilike(f'%{search}%'),
            )
        )
    if status == 'votou':
        query = query.filter_by(votou=True)
    elif status == 'pendente':
        query = query.filter_by(votou=False)
    if unidade_filtro:
        query = query.filter_by(unidade=unidade_filtro)
    if ativo_filtro == 'ativo':
        query = query.filter_by(ativo=True)
    elif ativo_filtro == 'inativo':
        query = query.filter_by(ativo=False)

    pagination = query.order_by(Funcionario.unidade, Funcionario.nome).paginate(
        page=page, per_page=25, error_out=False
    )
    return render_template(
        'admin/funcionarios.html',
        funcionarios=pagination.items,
        pagination=pagination,
        unidades=UNIDADES,
        search=search,
        status=status,
        unidade_filtro=unidade_filtro,
        ativo_filtro=ativo_filtro,
        total_geral=Funcionario.query.count(),
        total_votou=Funcionario.query.filter_by(votou=True).count(),
        total_pendente=Funcionario.query.filter_by(votou=False).count(),
        total_ativo=Funcionario.query.filter_by(ativo=True).count(),
        total_inativo=Funcionario.query.filter_by(ativo=False).count(),
    )


@admin_bp.route('/funcionarios/adicionar', methods=['POST'])
@login_required
def adicionar_funcionario():
    matricula = request.form['matricula'].strip()
    nome = request.form['nome'].strip()
    unidade = request.form['unidade'].strip()
    if matricula and nome and unidade:
        if Funcionario.query.filter_by(matricula=matricula).first():
            flash('Matrícula já cadastrada.', 'danger')
        else:
            f = Funcionario(
                matricula=matricula,
                nome=nome,
                unidade=unidade,
                setor=request.form.get('setor', '').strip() or None,
                cargo=request.form.get('cargo', '').strip() or None,
                data_admissao=_parse_date(request.form.get('data_admissao')),
                data_nascimento=_parse_date(request.form.get('data_nascimento')),
                ativo=request.form.get('ativo', '1') != '0',
            )
            db.session.add(f)
            db.session.commit()
            flash(f'Funcionário adicionado em {unidade}!', 'success')
    return redirect(url_for('admin.funcionarios'))


@admin_bp.route('/funcionarios/editar/<int:id>', methods=['POST'])
@login_required
def editar_funcionario(id):
    f = Funcionario.query.get_or_404(id)
    nova_matricula = request.form['matricula'].strip()
    # Verifica duplicata apenas se a matrícula mudou
    if nova_matricula != f.matricula and Funcionario.query.filter_by(matricula=nova_matricula).first():
        flash('Matrícula já cadastrada para outro funcionário.', 'danger')
        return redirect(url_for('admin.funcionarios'))
    f.matricula = nova_matricula
    f.nome = request.form['nome'].strip()
    f.unidade = request.form['unidade'].strip()
    f.setor = request.form.get('setor', '').strip() or None
    f.cargo = request.form.get('cargo', '').strip() or None
    f.data_admissao = _parse_date(request.form.get('data_admissao'))
    f.data_nascimento = _parse_date(request.form.get('data_nascimento'))
    f.ativo = request.form.get('ativo', '1') != '0'
    db.session.commit()
    flash(f'Funcionário {f.nome} atualizado!', 'success')
    return redirect(url_for('admin.funcionarios'))


@admin_bp.route('/funcionarios/modelo')
@login_required
def modelo_funcionarios():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Funcionarios'
    ws.append(['matricula', 'nome', 'unidade', 'setor', 'cargo', 'data_admissao', 'data_nascimento'])
    ws.append(['00001', 'João da Silva', 'UTD SOBRADINHO', 'Operações', 'Eletricista', '2015-03-20', '1985-07-10'])
    ws.append(['00002', 'Maria Oliveira', 'UTD PLANALTINA', 'Administrativo', 'Analista', '2018-08-15', '1990-02-22'])

    col_widths = {'A': 14, 'B': 40, 'C': 30, 'D': 20, 'E': 20, 'F': 16, 'G': 16}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='modelo_funcionarios.xlsx',
    )


@admin_bp.route('/funcionarios/importar', methods=['POST'])
@login_required
def importar_funcionarios():
    arquivo = request.files.get('planilha')
    if not arquivo or not arquivo.filename.endswith('.xlsx'):
        flash('Envie um arquivo Excel (.xlsx) válido.', 'danger')
        return redirect(url_for('admin.funcionarios'))

    try:
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active
        existentes = {f.matricula: f for f in Funcionario.query.all()}
        unidades_validas = set(UNIDADES)
        adicionados = atualizados = erros = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            if len(row) < 3:
                erros += 1
                continue

            matricula = str(row[0]).strip() if row[0] is not None else ''
            nome = str(row[1]).strip() if row[1] is not None else ''
            unidade = str(row[2]).strip() if row[2] is not None else ''
            setor = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            cargo = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None
            data_admissao = _parse_date(row[5]) if len(row) > 5 and row[5] is not None else None
            data_nascimento = _parse_date(row[6]) if len(row) > 6 and row[6] is not None else None

            if not matricula or not nome or not unidade:
                erros += 1
                continue
            if unidade not in unidades_validas:
                erros += 1
                continue

            if matricula in existentes:
                f = existentes[matricula]
                f.nome = nome
                f.unidade = unidade
                if setor:
                    f.setor = setor
                if cargo:
                    f.cargo = cargo
                if data_admissao:
                    f.data_admissao = data_admissao
                if data_nascimento:
                    f.data_nascimento = data_nascimento
                atualizados += 1
            else:
                novo = Funcionario(
                    matricula=matricula, nome=nome, unidade=unidade,
                    setor=setor, cargo=cargo,
                    data_admissao=data_admissao, data_nascimento=data_nascimento,
                )
                db.session.add(novo)
                existentes[matricula] = novo
                adicionados += 1

        db.session.commit()
        wb.close()

        partes = []
        if adicionados:
            partes.append(f'{adicionados} adicionado(s)')
        if atualizados:
            partes.append(f'{atualizados} atualizado(s)')
        if erros:
            partes.append(f'{erros} linha(s) ignorada(s) por dado inválido')

        flash('Importação concluída: ' + ', '.join(partes) + '.', 'success' if not erros or adicionados or atualizados else 'warning')

    except Exception:
        db.session.rollback()
        flash('Erro ao processar a planilha. Verifique se o arquivo está no formato correto.', 'danger')

    return redirect(url_for('admin.funcionarios'))


@admin_bp.route('/funcionarios/remover/<int:id>')
@login_required
def remover_funcionario(id):
    funcionario = Funcionario.query.get_or_404(id)
    db.session.delete(funcionario)
    db.session.commit()
    flash('Funcionário removido.', 'warning')
    return redirect(url_for('admin.funcionarios'))

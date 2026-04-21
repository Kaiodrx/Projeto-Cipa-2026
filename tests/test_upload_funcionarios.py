"""
Testes de importação de funcionários via planilha Excel.
Cobre: upload válido, atualização de existentes, mistos, arquivo inválido,
       linha incompleta, download do modelo e proteção de rota.
"""
import io
import pytest
from openpyxl import Workbook
from app.models import Funcionario
from app.constants import UNIDADES

UNIDADE = UNIDADES[0]


# ── Utilitário ────────────────────────────────────────────────────────────────

def _fazer_xlsx(*linhas):
    """Cria um arquivo Excel em memória com as linhas informadas."""
    wb = Workbook()
    ws = wb.active
    ws.append(['matricula', 'nome', 'unidade'])
    for linha in linhas:
        ws.append(list(linha))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _upload(client, buf, nome_arquivo='funcionarios.xlsx'):
    return client.post(
        '/admin/funcionarios/importar',
        data={'planilha': (buf, nome_arquivo)},
        content_type='multipart/form-data',
        follow_redirects=True,
    )


# ── Upload válido ─────────────────────────────────────────────────────────────

def test_importar_cria_funcionarios(logged_in_client, db):
    buf = _fazer_xlsx(
        ('MAT100', 'Alice Souza', UNIDADE),
        ('MAT101', 'Bruno Lima', UNIDADE),
    )
    resp = _upload(logged_in_client, buf)

    assert resp.status_code == 200
    assert Funcionario.query.filter_by(matricula='MAT100').first() is not None
    assert Funcionario.query.filter_by(matricula='MAT101').first() is not None


def test_importar_exibe_mensagem_de_sucesso(logged_in_client, db):
    buf = _fazer_xlsx(('MAT200', 'Carlos Melo', UNIDADE))
    resp = _upload(logged_in_client, buf)

    assert 'concluída' in resp.data.decode().lower()


# ── Atualização de existentes ─────────────────────────────────────────────────

def test_importar_atualiza_funcionario_existente(logged_in_client, db, funcionario):
    # funcionario fixture: matricula='MAT001', nome='Maria Santos'
    buf = _fazer_xlsx(('MAT001', 'Maria Santos Atualizada', UNIDADES[1]))
    _upload(logged_in_client, buf)

    atualizado = Funcionario.query.filter_by(matricula='MAT001').first()
    assert atualizado.nome == 'Maria Santos Atualizada'
    assert atualizado.unidade == UNIDADES[1]


def test_importar_nao_duplica_funcionario_existente(logged_in_client, db, funcionario):
    buf = _fazer_xlsx(('MAT001', 'Maria Santos Atualizada', UNIDADE))
    _upload(logged_in_client, buf)

    total = Funcionario.query.filter_by(matricula='MAT001').count()
    assert total == 1


# ── Planilha mista (novos + existentes) ───────────────────────────────────────

def test_importar_misto_adiciona_e_atualiza(logged_in_client, db, funcionario):
    buf = _fazer_xlsx(
        ('MAT001', 'Maria Atualizada', UNIDADE),   # já existe
        ('MAT300', 'Novo Funcionario', UNIDADE),    # novo
    )
    _upload(logged_in_client, buf)

    assert Funcionario.query.filter_by(matricula='MAT001').first().nome == 'Maria Atualizada'
    assert Funcionario.query.filter_by(matricula='MAT300').first() is not None


# ── Arquivo inválido ──────────────────────────────────────────────────────────

def test_importar_arquivo_csv_e_rejeitado(logged_in_client, db):
    buf = io.BytesIO(b'matricula,nome,unidade\nMAT400,Joao,Unidade A')
    resp = _upload(logged_in_client, buf, nome_arquivo='funcionarios.csv')

    assert resp.status_code == 200
    assert Funcionario.query.filter_by(matricula='MAT400').first() is None


def test_importar_arquivo_invalido_exibe_mensagem_de_erro(logged_in_client, db):
    buf = io.BytesIO(b'matricula,nome,unidade\nMAT400,Joao,Unidade A')
    resp = _upload(logged_in_client, buf, nome_arquivo='funcionarios.csv')

    assert 'válido' in resp.data.decode().lower()


# ── Linhas incompletas ────────────────────────────────────────────────────────

def test_importar_linha_sem_nome_e_ignorada(logged_in_client, db):
    buf = _fazer_xlsx(('MAT500', '', UNIDADE))
    _upload(logged_in_client, buf)

    assert Funcionario.query.filter_by(matricula='MAT500').first() is None


def test_importar_linha_sem_matricula_e_ignorada(logged_in_client, db):
    buf = _fazer_xlsx(('', 'Sem Matricula', UNIDADE))
    _upload(logged_in_client, buf)

    assert Funcionario.query.filter_by(nome='Sem Matricula').first() is None


# ── Download do modelo ────────────────────────────────────────────────────────

def test_baixar_modelo_retorna_xlsx(logged_in_client, db):
    resp = logged_in_client.get('/admin/funcionarios/modelo')

    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.content_type


def test_baixar_modelo_e_arquivo_valido(logged_in_client, db):
    resp = logged_in_client.get('/admin/funcionarios/modelo')
    buf = io.BytesIO(resp.data)
    wb = Workbook()
    from openpyxl import load_workbook
    wb = load_workbook(buf)
    ws = wb.active
    cabecalho = [cell.value for cell in ws[1]]

    assert cabecalho == ['matricula', 'nome', 'unidade']


# ── Proteção de rota ──────────────────────────────────────────────────────────

def test_importar_sem_autenticacao_redireciona(client, db):
    buf = _fazer_xlsx(('MAT600', 'Invasor', UNIDADE))
    resp = _upload(client, buf)

    assert 'Acesso Administrativo' in resp.data.decode()
    assert Funcionario.query.filter_by(matricula='MAT600').first() is None


def test_baixar_modelo_sem_autenticacao_redireciona(client, db):
    resp = client.get('/admin/funcionarios/modelo', follow_redirects=True)

    assert 'Acesso Administrativo' in resp.data.decode()
